#!/usr/bin/env python3
"""
PDF Pre-Scan Tool

Analyzes PDF files to detect characteristics that may cause Adobe API failures.
Useful for identifying "risky" PDFs before uploading to the processing pipeline.

Risk factors detected:
- Large page dimensions (newspapers, posters, scanned docs)
- Mixed page sizes within document
- High image-to-page ratio (image-heavy documents)
- High file density (MB per page)

Usage:
    ./bin/prescan-pdf.py <pdf_file_or_directory> [options]

Examples:
    ./bin/prescan-pdf.py document.pdf
    ./bin/prescan-pdf.py /path/to/pdfs/
    ./bin/prescan-pdf.py document.pdf --json
    ./bin/prescan-pdf.py /path/to/pdfs/ --risky-only
    ./bin/prescan-pdf.py /path/to/pdfs/ --xlsx report.xlsx
"""

import argparse
import io
import json
import os
import sys
from pathlib import Path


def prescan_pdf_for_risk(pdf_path: str) -> dict:
    """
    Pre-scan a PDF to detect characteristics that may cause Adobe API failures.
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        dict with risk assessment results
    """
    from pypdf import PdfReader
    
    filename = os.path.basename(pdf_path)
    file_size = os.path.getsize(pdf_path)
    file_size_mb = file_size / (1024 * 1024)
    
    result = {
        'filename': filename,
        'path': pdf_path,
        'file_size_mb': round(file_size_mb, 2),
        'is_risky': False,
        'risk_factors': [],
        'recommended_pages_per_chunk': 90,
        'page_count': 0,
        'has_large_pages': False,
        'has_mixed_sizes': False,
        'is_image_heavy': False,
        'analysis_details': {}
    }
    
    try:
        with open(pdf_path, 'rb') as f:
            pdf_content = f.read()
        
        reader = PdfReader(io.BytesIO(pdf_content))
        num_pages = len(reader.pages)
        result['page_count'] = num_pages
        
        if num_pages == 0:
            result['risk_factors'].append('Empty PDF (0 pages)')
            result['is_risky'] = True
            return result
        
        # Analyze page dimensions
        page_sizes = set()
        large_page_count = 0
        total_images = 0
        page_details = []
        
        # Standard page sizes in points (72 points = 1 inch)
        LARGE_DIMENSION_THRESHOLD = 1200  # ~16.7 inches - larger than tabloid
        
        for page_num, page in enumerate(reader.pages):
            mediabox = page.mediabox
            width = float(mediabox.width)
            height = float(mediabox.height)
            
            # Normalize to portrait orientation for comparison
            w, h = (min(width, height), max(width, height))
            page_sizes.add((round(w), round(h)))
            
            # Check for large/unusual dimensions
            is_large = width > LARGE_DIMENSION_THRESHOLD or height > LARGE_DIMENSION_THRESHOLD
            if is_large:
                large_page_count += 1
            
            # Count images on page
            page_images = 0
            if '/XObject' in page.get('/Resources', {}):
                xobjects = page['/Resources'].get('/XObject', {})
                if hasattr(xobjects, 'keys'):
                    for obj_name in xobjects.keys():
                        try:
                            xobj = xobjects[obj_name]
                            if xobj.get('/Subtype') == '/Image':
                                page_images += 1
                                total_images += 1
                        except:
                            pass
            
            page_details.append({
                'page': page_num + 1,
                'width': round(width, 1),
                'height': round(height, 1),
                'is_large': is_large,
                'images': page_images
            })
        
        # Assess risk factors
        
        # 1. Large/unusual page dimensions
        if large_page_count > 0:
            result['has_large_pages'] = True
            result['risk_factors'].append(f'{large_page_count} pages with large dimensions (>{LARGE_DIMENSION_THRESHOLD}pt)')
            result['is_risky'] = True
        
        # 2. Mixed page sizes
        if len(page_sizes) > 2:
            result['has_mixed_sizes'] = True
            result['risk_factors'].append(f'{len(page_sizes)} different page sizes detected')
            result['is_risky'] = True
        
        # 3. Image-heavy documents
        images_per_page = total_images / num_pages if num_pages > 0 else 0
        if images_per_page >= 1.0 or total_images > num_pages * 0.8:
            result['is_image_heavy'] = True
            result['risk_factors'].append(f'Image-heavy ({total_images} images in {num_pages} pages, {images_per_page:.1f}/page)')
            result['is_risky'] = True
        
        # 4. File size per page
        mb_per_page = file_size_mb / num_pages if num_pages > 0 else 0
        if mb_per_page > 2.0:
            result['risk_factors'].append(f'High density ({mb_per_page:.1f} MB/page)')
            result['is_risky'] = True
        
        # 5. Total file size check
        if file_size_mb > 100:
            result['risk_factors'].append(f'Exceeds Adobe API limit (100MB max, file is {file_size_mb:.1f}MB)')
            result['is_risky'] = True
        
        # Determine recommended chunking
        if result['is_risky']:
            if num_pages <= 10:
                result['recommended_pages_per_chunk'] = 1
            elif num_pages <= 20:
                result['recommended_pages_per_chunk'] = 2
            else:
                result['recommended_pages_per_chunk'] = 5
        
        # Store analysis details
        result['analysis_details'] = {
            'page_sizes_found': [list(ps) for ps in page_sizes],
            'large_page_count': large_page_count,
            'total_images': total_images,
            'images_per_page': round(images_per_page, 2),
            'mb_per_page': round(mb_per_page, 2),
            'page_details': page_details[:10] if len(page_details) > 10 else page_details  # Limit for readability
        }
        
        if len(page_details) > 10:
            result['analysis_details']['note'] = f'Showing first 10 of {len(page_details)} pages'
        
    except Exception as e:
        result['error'] = str(e)
        result['is_risky'] = True
        result['risk_factors'].append(f'Error reading PDF: {e}')
    
    return result


def print_report(result: dict, verbose: bool = False):
    """Print a human-readable report."""
    print(f"\n{'='*60}")
    print(f"PDF Pre-Scan Report: {result['filename']}")
    print(f"{'='*60}")
    print(f"  Path: {result['path']}")
    print(f"  File size: {result['file_size_mb']:.2f} MB")
    print(f"  Pages: {result['page_count']}")
    
    if result['page_count'] > 0:
        print(f"  MB per page: {result['analysis_details'].get('mb_per_page', 'N/A')}")
        print(f"  Images detected: {result['analysis_details'].get('total_images', 0)}")
        print(f"  Unique page sizes: {len(result['analysis_details'].get('page_sizes_found', []))}")
    
    print()
    if result['is_risky']:
        print(f"  ⚠️  RISK ASSESSMENT: RISKY")
        print(f"  Risk factors:")
        for factor in result['risk_factors']:
            print(f"    - {factor}")
    else:
        print(f"  ✅ RISK ASSESSMENT: NORMAL")
    
    print()
    print(f"  Recommended pages per chunk: {result['recommended_pages_per_chunk']}")
    
    if result['page_count'] > 0:
        expected_chunks = max(1, result['page_count'] // result['recommended_pages_per_chunk'])
        print(f"  Expected chunks: ~{expected_chunks}")
    
    if verbose and 'page_details' in result.get('analysis_details', {}):
        print()
        print("  Page details:")
        for pd in result['analysis_details']['page_details']:
            large_marker = " [LARGE]" if pd['is_large'] else ""
            print(f"    Page {pd['page']}: {pd['width']}x{pd['height']}pt, {pd['images']} images{large_marker}")
        if 'note' in result['analysis_details']:
            print(f"    ({result['analysis_details']['note']})")
    
    print()


def write_excel_report(results: list, output_path: str):
    """Write results to an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Pre-Scan Results"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    risky_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    normal_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Filename",
        "Path",
        "File Size (MB)",
        "Pages",
        "MB/Page",
        "Images",
        "Images/Page",
        "Unique Page Sizes",
        "Large Pages",
        "Risk Assessment",
        "Risk Factors",
        "Recommended Pages/Chunk",
        "Expected Chunks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row_num, result in enumerate(results, 2):
        analysis = result.get('analysis_details', {})
        page_count = result['page_count']
        recommended = result['recommended_pages_per_chunk']
        expected_chunks = max(1, page_count // recommended) if page_count > 0 else 0
        
        # Use newlines instead of semicolons for risk factors (for text wrapping)
        risk_factors_text = "\n".join(result['risk_factors']) if result['risk_factors'] else ""
        
        row_data = [
            result['filename'],
            result['path'],
            result['file_size_mb'],
            page_count,
            analysis.get('mb_per_page', 0),
            analysis.get('total_images', 0),
            analysis.get('images_per_page', 0),
            len(analysis.get('page_sizes_found', [])),
            analysis.get('large_page_count', 0),
            "RISKY" if result['is_risky'] else "NORMAL",
            risk_factors_text,
            recommended,
            expected_chunks
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Apply risk-based coloring to the Risk Assessment column
            if col == 10:  # Risk Assessment column
                if result['is_risky']:
                    cell.fill = risky_fill
                else:
                    cell.fill = normal_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Apply text wrapping to Risk Factors column
            if col == 11:  # Risk Factors column
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        # Skip Risk Factors column - set fixed width for wrapped text
        if col == 11:
            ws.column_dimensions[get_column_letter(col)].width = 40
            continue
            
        max_length = len(headers[col - 1])
        for row in range(2, len(results) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Cap width and add padding
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet if multiple files
    if len(results) > 1:
        summary_ws = wb.create_sheet("Summary")
        
        total_files = len(results)
        risky_count = sum(1 for r in results if r['is_risky'])
        normal_count = total_files - risky_count
        total_pages = sum(r['page_count'] for r in results)
        total_size = sum(r['file_size_mb'] for r in results)
        
        summary_data = [
            ["Metric", "Value"],
            ["Total PDFs Scanned", total_files],
            ["Risky PDFs", risky_count],
            ["Normal PDFs", normal_count],
            ["Total Pages", total_pages],
            ["Total Size (MB)", round(total_size, 2)],
            ["Risky Percentage", f"{(risky_count/total_files*100):.1f}%" if total_files > 0 else "0%"]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if row_num == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 15
    
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Pre-scan PDF files to detect characteristics that may cause Adobe API failures.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s document.pdf                    # Scan a single PDF
  %(prog)s /path/to/pdfs/                  # Scan all PDFs in a directory
  %(prog)s document.pdf --json             # Output as JSON
  %(prog)s /path/to/pdfs/ --risky-only     # Only show risky PDFs
  %(prog)s document.pdf -v                 # Verbose output with page details
  %(prog)s /path/to/pdfs/ --xlsx report.xlsx  # Export to Excel
        """
    )
    parser.add_argument('path', help='PDF file or directory containing PDFs')
    parser.add_argument('--json', action='store_true', help='Output results as JSON')
    parser.add_argument('--xlsx', metavar='FILE', help='Export results to Excel spreadsheet')
    parser.add_argument('--risky-only', action='store_true', help='Only show risky PDFs')
    parser.add_argument('-v', '--verbose', action='store_true', help='Show detailed page information')
    parser.add_argument('--recursive', '-r', action='store_true', help='Recursively scan directories')
    
    args = parser.parse_args()
    
    # Check for pypdf
    try:
        from pypdf import PdfReader
    except ImportError:
        print("Error: pypdf is required. Install with: pip install pypdf", file=sys.stderr)
        sys.exit(1)
    
    # Collect PDF files to scan
    pdf_files = []
    path = Path(args.path)
    
    if path.is_file():
        if path.suffix.lower() == '.pdf':
            pdf_files.append(str(path))
        else:
            print(f"Error: {args.path} is not a PDF file", file=sys.stderr)
            sys.exit(1)
    elif path.is_dir():
        pattern = '**/*.pdf' if args.recursive else '*.pdf'
        pdf_files = [str(p) for p in path.glob(pattern)]
        if not pdf_files:
            print(f"No PDF files found in {args.path}", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"Error: {args.path} does not exist", file=sys.stderr)
        sys.exit(1)
    
    # Scan PDFs
    results = []
    for pdf_path in sorted(pdf_files):
        result = prescan_pdf_for_risk(pdf_path)
        
        if args.risky_only and not result['is_risky']:
            continue
        
        results.append(result)
    
    # Output results
    if args.xlsx:
        # Check for openpyxl
        try:
            from openpyxl import Workbook
        except ImportError:
            print("Error: openpyxl is required for Excel export. Install with: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        
        if not results:
            print("No PDFs to export.", file=sys.stderr)
            sys.exit(1)
        
        write_excel_report(results, args.xlsx)
        
        # Also print summary to console
        risky_count = sum(1 for r in results if r['is_risky'])
        print(f"Scanned {len(results)} PDFs, {risky_count} risky")
    elif args.json:
        print(json.dumps(results, indent=2))
    else:
        if not results:
            if args.risky_only:
                print("No risky PDFs found.")
            else:
                print("No PDFs to report.")
        else:
            for result in results:
                print_report(result, verbose=args.verbose)
            
            # Summary for multiple files
            if len(results) > 1:
                risky_count = sum(1 for r in results if r['is_risky'])
                print(f"\n{'='*60}")
                print(f"SUMMARY: {len(results)} PDFs scanned, {risky_count} risky")
                print(f"{'='*60}")


if __name__ == '__main__':
    main()
