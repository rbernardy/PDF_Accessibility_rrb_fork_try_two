#!/usr/bin/env python3
"""
Export DynamoDB table records to an Excel .xlsx file.

Usage:
    python export-dynamodb-to-excel.py --table-name MyTable
    python export-dynamodb-to-excel.py --table-name MyTable --max-records 100 --sort-by created_at --sort-order desc
"""

import argparse
import re
import sys
from datetime import datetime
from decimal import Decimal

# Regex pattern for illegal XML characters that Excel doesn't support
ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)

# Check for required dependencies
try:
    import boto3
except ImportError:
    print("Error: boto3 is required. Install it with: pip install boto3")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def sanitize_for_excel(value):
    """Remove illegal characters that Excel doesn't support."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def parse_dynamodb_value(value):
    """Convert DynamoDB typed value to Python native type."""
    if value is None:
        return None
    
    if isinstance(value, dict):
        if len(value) == 1:
            type_key = list(value.keys())[0]
            val = value[type_key]
            
            if type_key == 'S':  # String
                return val
            elif type_key == 'N':  # Number
                return float(val) if '.' in val else int(val)
            elif type_key == 'BOOL':  # Boolean
                return val
            elif type_key == 'NULL':  # Null
                return None
            elif type_key == 'B':  # Binary
                return f"<binary: {len(val)} bytes>"
            elif type_key == 'SS':  # String Set
                return ', '.join(sorted(val))
            elif type_key == 'NS':  # Number Set
                return ', '.join(sorted(val, key=lambda x: float(x)))
            elif type_key == 'BS':  # Binary Set
                return f"<binary set: {len(val)} items>"
            elif type_key == 'L':  # List
                parsed_items = [parse_dynamodb_value(item) for item in val]
                return str(parsed_items)
            elif type_key == 'M':  # Map
                parsed_map = {k: parse_dynamodb_value(v) for k, v in val.items()}
                return str(parsed_map)
        
        # If it's a map without type wrapper (from resource interface)
        return {k: parse_dynamodb_value(v) for k, v in value.items()}
    
    # Handle boto3 resource interface types (already deserialized)
    if isinstance(value, Decimal):
        return float(value) if value % 1 else int(value)
    elif isinstance(value, set):
        return ', '.join(str(v) for v in sorted(value))
    elif isinstance(value, list):
        return str([parse_dynamodb_value(v) for v in value])
    
    return value


def flatten_item(item):
    """Flatten a DynamoDB item to simple key-value pairs."""
    flattened = {}
    for key, value in item.items():
        flattened[key] = parse_dynamodb_value(value)
    return flattened


def scan_table(table_name, max_records=None):
    """Scan DynamoDB table and return all records."""
    dynamodb = boto3.resource('dynamodb')
    table = dynamodb.Table(table_name)
    
    items = []
    scan_kwargs = {}
    total_scanned = 0
    
    print(f"Scanning table '{table_name}'...")
    
    while True:
        response = table.scan(**scan_kwargs)
        batch_items = response.get('Items', [])
        total_scanned += response.get('ScannedCount', 0)
        
        for item in batch_items:
            items.append(flatten_item(item))
            
            if max_records and len(items) >= max_records:
                print(f"  Reached max records limit: {max_records}")
                return items
        
        print(f"  Scanned {total_scanned} records, collected {len(items)} items...")
        
        # Check for pagination
        if 'LastEvaluatedKey' in response:
            scan_kwargs['ExclusiveStartKey'] = response['LastEvaluatedKey']
        else:
            break
    
    return items


def get_all_columns(items):
    """Extract all unique column names from items."""
    columns = set()
    for item in items:
        columns.update(item.keys())
    return sorted(columns)


def sort_items(items, sort_by, sort_order):
    """Sort items by specified column."""
    if not sort_by:
        return items
    
    reverse = sort_order.lower() == 'desc'
    
    def sort_key(item):
        value = item.get(sort_by)
        if value is None:
            return (1, '')  # Put None values at the end
        return (0, value)
    
    try:
        return sorted(items, key=sort_key, reverse=reverse)
    except TypeError:
        # If comparison fails, convert to string
        def string_sort_key(item):
            value = item.get(sort_by)
            if value is None:
                return (1, '')
            return (0, str(value))
        return sorted(items, key=string_sort_key, reverse=reverse)


def create_excel(items, columns, output_file):
    """Create Excel file from items."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DynamoDB Export"
    
    # Write header row with bold font
    header_font = Font(bold=True)
    for col_idx, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
    
    # Write data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, column_name in enumerate(columns, 1):
            value = item.get(column_name, '')
            # Convert complex types to string for Excel
            if isinstance(value, (dict, list)):
                value = str(value)
            # Sanitize value to remove illegal XML characters
            value = sanitize_for_excel(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns based on content
    for col_idx, column_name in enumerate(columns, 1):
        max_length = len(str(column_name))
        
        for row_idx in range(2, min(len(items) + 2, 102)):  # Sample first 100 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 50))
        
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Create failure summary sheet
    unique_pdf_keys = create_failure_summary_sheet(wb, items)
    
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")
    
    return unique_pdf_keys


def create_failure_summary_sheet(wb, items):
    """Create a summary sheet counting failures per pdf_key."""
    ws = wb.create_sheet(title="Failure Summary")
    
    # Count failures per pdf_key
    failure_counts = {}
    for item in items:
        pdf_key = item.get('pdf_key', '')
        if pdf_key:
            failure_counts[pdf_key] = failure_counts.get(pdf_key, 0) + 1
    
    # Sort by pdf_key
    sorted_keys = sorted(failure_counts.keys())
    
    # Write headers
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="pdf_key").font = header_font
    ws.cell(row=1, column=2, value="failure_count").font = header_font
    
    # Write data
    for row_idx, pdf_key in enumerate(sorted_keys, 2):
        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(pdf_key))
        ws.cell(row=row_idx, column=2, value=failure_counts[pdf_key])
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 15
    
    return len(sorted_keys)



def main():
    parser = argparse.ArgumentParser(
        description='Export DynamoDB table records to an Excel .xlsx file.'
    )
    parser.add_argument(
        '--table-name',
        required=True,
        help='DynamoDB table name'
    )
    parser.add_argument(
        '--max-records',
        type=int,
        default=None,
        help='Maximum number of records to export (default: all)'
    )
    parser.add_argument(
        '--sort-by',
        default=None,
        help='Column name to sort by'
    )
    parser.add_argument(
        '--sort-order',
        choices=['asc', 'desc'],
        default='asc',
        help="Sort order: 'asc' or 'desc' (default: asc)"
    )
    parser.add_argument(
        '--output',
        default=None,
        help='Output filename (default: {table-name}_export_{timestamp}.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Generate default output filename if not provided
    if args.output:
        output_file = args.output
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{args.table_name}_export_{timestamp}.xlsx"
    
    try:
        # Scan the table
        items = scan_table(args.table_name, args.max_records)
        
        if not items:
            print("No records found in the table.")
            return
        
        # Get all columns
        columns = get_all_columns(items)
        print(f"Found {len(columns)} columns: {', '.join(columns)}")
        
        # Sort if requested
        if args.sort_by:
            if args.sort_by not in columns:
                print(f"Warning: Sort column '{args.sort_by}' not found in data. Skipping sort.")
            else:
                print(f"Sorting by '{args.sort_by}' ({args.sort_order})...")
                items = sort_items(items, args.sort_by, args.sort_order)
        
        # Create Excel file
        print(f"Creating Excel file with {len(items)} records...")
        unique_pdf_keys = create_excel(items, columns, output_file)
        
        # Print summary
        print("\n--- Export Summary ---")
        print(f"Table: {args.table_name}")
        print(f"Records exported: {len(items)}")
        print(f"Columns: {len(columns)}")
        print(f"Unique pdf_keys: {unique_pdf_keys}")
        print(f"Output file: {output_file}")
        
    except boto3.client('dynamodb').exceptions.ResourceNotFoundException:
        print(f"Error: Table '{args.table_name}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
