#!/usr/bin/env python3
"""
S3 Bucket Listing to Excel

Lists objects in an S3 bucket and exports the data to an Excel spreadsheet.

Usage:
    ./bin/s3-listing-to-xlsx.py <bucket-name> [options]

Examples:
    ./bin/s3-listing-to-xlsx.py my-bucket
    ./bin/s3-listing-to-xlsx.py my-bucket --prefix pdf/
    ./bin/s3-listing-to-xlsx.py my-bucket --output listing.xlsx
    ./bin/s3-listing-to-xlsx.py my-bucket --prefix queue/ --output queue-files.xlsx
"""

import argparse
import json
import sys
from datetime import datetime

import boto3
from botocore.exceptions import ClientError


def list_s3_objects(bucket: str, prefix: str = None) -> list:
    """
    List all objects in an S3 bucket.
    
    Args:
        bucket: S3 bucket name
        prefix: Optional prefix to filter objects
        
    Returns:
        List of object metadata dicts
    """
    s3 = boto3.client('s3')
    objects = []
    
    paginator = s3.get_paginator('list_objects_v2')
    
    params = {'Bucket': bucket}
    if prefix:
        params['Prefix'] = prefix
    
    try:
        for page in paginator.paginate(**params):
            if 'Contents' in page:
                for obj in page['Contents']:
                    objects.append({
                        'Key': obj['Key'],
                        'Size': obj['Size'],
                        'LastModified': obj['LastModified'],
                        'ETag': obj.get('ETag', '').strip('"'),
                        'StorageClass': obj.get('StorageClass', 'STANDARD')
                    })
    except ClientError as e:
        print(f"Error listing bucket: {e}", file=sys.stderr)
        sys.exit(1)
    
    return objects


def format_size(size_bytes: int) -> str:
    """Format bytes as human-readable size."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"


def write_excel_report(objects: list, bucket: str, prefix: str, output_path: str):
    """Write S3 listing to an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "S3 Objects"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Key",
        "Folder",
        "Filename",
        "Extension",
        "Size (Bytes)",
        "Size (Human)",
        "Last Modified",
        "Storage Class",
        "ETag"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row_num, obj in enumerate(objects, 2):
        key = obj['Key']
        
        # Parse key into folder and filename
        if '/' in key:
            parts = key.rsplit('/', 1)
            folder = parts[0] + '/'
            filename = parts[1]
        else:
            folder = ''
            filename = key
        
        # Get extension
        if '.' in filename:
            extension = filename.rsplit('.', 1)[1].lower()
        else:
            extension = ''
        
        # Format last modified
        last_modified = obj['LastModified']
        if isinstance(last_modified, datetime):
            last_modified_str = last_modified.strftime('%Y-%m-%d %H:%M:%S')
        else:
            last_modified_str = str(last_modified)
        
        row_data = [
            key,
            folder,
            filename,
            extension,
            obj['Size'],
            format_size(obj['Size']),
            last_modified_str,
            obj['StorageClass'],
            obj['ETag']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Right-align size columns
            if col in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column widths
    column_widths = [60, 40, 30, 10, 15, 12, 20, 15, 35]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    total_size = sum(obj['Size'] for obj in objects)
    total_files = len(objects)
    
    # Count by extension
    ext_counts = {}
    ext_sizes = {}
    for obj in objects:
        key = obj['Key']
        if '.' in key:
            ext = key.rsplit('.', 1)[1].lower()
        else:
            ext = '(no extension)'
        ext_counts[ext] = ext_counts.get(ext, 0) + 1
        ext_sizes[ext] = ext_sizes.get(ext, 0) + obj['Size']
    
    # Count by folder (top-level)
    folder_counts = {}
    folder_sizes = {}
    for obj in objects:
        key = obj['Key']
        if '/' in key:
            top_folder = key.split('/')[0] + '/'
        else:
            top_folder = '(root)'
        folder_counts[top_folder] = folder_counts.get(top_folder, 0) + 1
        folder_sizes[top_folder] = folder_sizes.get(top_folder, 0) + obj['Size']
    
    # Write summary
    summary_data = [
        ["S3 Bucket Listing Summary", ""],
        ["", ""],
        ["Bucket", bucket],
        ["Prefix", prefix or "(none)"],
        ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["Total Objects", total_files],
        ["Total Size", format_size(total_size)],
        ["Total Size (Bytes)", total_size],
        ["", ""],
        ["By Extension", "Count", "Size"],
    ]
    
    for ext in sorted(ext_counts.keys()):
        summary_data.append([ext, ext_counts[ext], format_size(ext_sizes[ext])])
    
    summary_data.append(["", ""])
    summary_data.append(["By Top-Level Folder", "Count", "Size"])
    
    for folder in sorted(folder_counts.keys()):
        summary_data.append([folder, folder_counts[folder], format_size(folder_sizes[folder])])
    
    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["By Extension", "By Top-Level Folder"]:
                cell.font = Font(bold=True)
    
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 15
    
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='List S3 bucket objects and export to Excel.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s my-bucket                           # List entire bucket
  %(prog)s my-bucket --prefix pdf/             # List only pdf/ folder
  %(prog)s my-bucket -o listing.xlsx           # Custom output filename
  %(prog)s my-bucket --prefix queue/ -o q.xlsx # Combine options
        """
    )
    parser.add_argument('bucket', help='S3 bucket name')
    parser.add_argument('--prefix', '-p', help='Filter by key prefix (e.g., "pdf/")')
    parser.add_argument('--output', '-o', help='Output Excel filename (default: <bucket>_listing.xlsx)')
    parser.add_argument('--json', action='store_true', help='Also output JSON to stdout')
    
    args = parser.parse_args()
    
    # Check for openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    
    # Determine output filename
    if args.output:
        output_path = args.output
    else:
        safe_bucket = args.bucket.replace('/', '_').replace('\\', '_')
        if args.prefix:
            safe_prefix = args.prefix.rstrip('/').replace('/', '_')
            output_path = f"{safe_bucket}_{safe_prefix}_listing.xlsx"
        else:
            output_path = f"{safe_bucket}_listing.xlsx"
    
    print(f"Listing s3://{args.bucket}/{args.prefix or ''}...")
    
    # List objects
    objects = list_s3_objects(args.bucket, args.prefix)
    
    if not objects:
        print("No objects found.")
        sys.exit(0)
    
    print(f"Found {len(objects)} objects")
    
    # Output JSON if requested
    if args.json:
        # Convert datetime to string for JSON serialization
        json_objects = []
        for obj in objects:
            json_obj = obj.copy()
            if isinstance(json_obj['LastModified'], datetime):
                json_obj['LastModified'] = json_obj['LastModified'].isoformat()
            json_objects.append(json_obj)
        print(json.dumps(json_objects, indent=2))
    
    # Write Excel
    write_excel_report(objects, args.bucket, args.prefix, output_path)
    
    # Print summary
    total_size = sum(obj['Size'] for obj in objects)
    print(f"Total size: {format_size(total_size)}")


if __name__ == '__main__':
    main()
