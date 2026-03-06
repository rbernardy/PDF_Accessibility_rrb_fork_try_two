#!/usr/bin/env python3
"""
Advanced PDF Pre-Scan Tool (PyMuPDF)

Deep analysis of PDF files using PyMuPDF to extract rich structural data.
Outputs detailed metrics to Excel for analysis of Adobe API failure patterns.

Metrics extracted:
- Image details: dimensions, colorspace, compression, bits per component
- Text analysis: character count, font count, text area ratio
- Page complexity: drawings, annotations, links
- Document structure: embedded files, JavaScript, forms
- Table detection (if pdfplumber available)

Usage:
    ./bin/prescan-pdf-advanced.py <pdf_file_or_directory> --xlsx output.xlsx

Examples:
    ./bin/prescan-pdf-advanced.py document.pdf --xlsx report.xlsx
    ./bin/prescan-pdf-advanced.py /path/to/pdfs/ --xlsx analysis.xlsx
    ./bin/prescan-pdf-advanced.py /path/to/pdfs/ -r --xlsx recursive_scan.xlsx
"""

import argparse
import io
import json
import os
import sys
from datetime import datetime
from pathlib import Path


def analyze_pdf_advanced(pdf_path: str) -> dict:
    """
    Deep analysis of a PDF using PyMuPDF.
    
    Returns comprehensive metrics about the PDF structure and content.
    """
    import fitz  # PyMuPDF
    
    filename = os.path.basename(pdf_path)
    file_size = os.path.getsize(pdf_path)
    file_size_mb = file_size / (1024 * 1024)
    
    result = {
        'filename': filename,
        'path': pdf_path,
        'file_size_bytes': file_size,
        'file_size_mb': round(file_size_mb, 2),
        'error': None,
        
        # Document-level metrics
        'page_count': 0,
        'pdf_version': '',
        'is_encrypted': False,
        'is_repaired': False,
        'has_javascript': False,
        'has_embedded_files': False,
        'has_forms': False,
        'has_annotations': False,
        'metadata': {},
        
        # Aggregate metrics
        'total_images': 0,
        'total_image_bytes': 0,
        'total_text_chars': 0,
        'total_fonts': 0,
        'total_drawings': 0,
        'total_links': 0,
        'total_annotations': 0,
        
        # Image analysis
        'image_colorspaces': {},
        'image_compressions': {},
        'max_image_width': 0,
        'max_image_height': 0,
        'max_image_pixels': 0,
        'avg_image_pixels': 0,
        'images_over_1mp': 0,
        'images_over_4mp': 0,
        
        # Page analysis
        'unique_page_sizes': 0,
        'page_sizes': [],
        'max_page_width_pt': 0,
        'max_page_height_pt': 0,
        'pages_over_tabloid': 0,
        
        # Complexity indicators
        'mb_per_page': 0,
        'images_per_page': 0,
        'chars_per_page': 0,
        'drawings_per_page': 0,
        
        # Per-page details (first 20 pages)
        'page_details': [],
        
        # Risk assessment
        'risk_factors': [],
        'complexity_score': 0,
    }
    
    try:
        doc = fitz.open(pdf_path)
        
        # Document-level info
        result['page_count'] = len(doc)
        result['pdf_version'] = f"{doc.metadata.get('format', 'Unknown')}"
        result['is_encrypted'] = doc.is_encrypted
        result['is_repaired'] = doc.is_repaired
        
        # Check for JavaScript
        try:
            js = doc.get_page_javascripts()
            result['has_javascript'] = bool(js)
        except:
            pass
        
        # Check for embedded files
        try:
            result['has_embedded_files'] = doc.embfile_count() > 0
        except:
            pass
        
        # Metadata
        result['metadata'] = {
            'title': doc.metadata.get('title', ''),
            'author': doc.metadata.get('author', ''),
            'subject': doc.metadata.get('subject', ''),
            'creator': doc.metadata.get('creator', ''),
            'producer': doc.metadata.get('producer', ''),
            'creation_date': doc.metadata.get('creationDate', ''),
            'mod_date': doc.metadata.get('modDate', ''),
        }
        
        # Collect all fonts used in document
        all_fonts = set()
        page_sizes_set = set()
        all_images = []
        
        # Analyze each page
        for page_num in range(len(doc)):
            page = doc[page_num]
            
            page_detail = {
                'page': page_num + 1,
                'width_pt': round(page.rect.width, 1),
                'height_pt': round(page.rect.height, 1),
                'rotation': page.rotation,
                'image_count': 0,
                'image_bytes': 0,
                'text_chars': 0,
                'font_count': 0,
                'drawing_count': 0,
                'link_count': 0,
                'annot_count': 0,
                'images': [],
            }
            
            # Page size
            w, h = page.rect.width, page.rect.height
            # Normalize to portrait
            norm_size = (round(min(w, h), 0), round(max(w, h), 0))
            page_sizes_set.add(norm_size)
            
            result['max_page_width_pt'] = max(result['max_page_width_pt'], w)
            result['max_page_height_pt'] = max(result['max_page_height_pt'], h)
            
            # Tabloid is 792x1224 points (11x17 inches)
            if w > 1224 or h > 1224:
                result['pages_over_tabloid'] += 1
            
            # Extract images
            try:
                image_list = page.get_images(full=True)
                page_detail['image_count'] = len(image_list)
                result['total_images'] += len(image_list)
                
                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    try:
                        base_image = doc.extract_image(xref)
                        if base_image:
                            img_width = base_image.get('width', 0)
                            img_height = base_image.get('height', 0)
                            img_colorspace = base_image.get('colorspace', 0)
                            img_bpc = base_image.get('bpc', 0)
                            img_size = len(base_image.get('image', b''))
                            
                            # Map colorspace number to name
                            cs_names = {1: 'Gray', 3: 'RGB', 4: 'CMYK'}
                            cs_name = cs_names.get(img_colorspace, f'CS{img_colorspace}')
                            
                            img_info = {
                                'width': img_width,
                                'height': img_height,
                                'pixels': img_width * img_height,
                                'colorspace': cs_name,
                                'bpc': img_bpc,
                                'size_bytes': img_size,
                            }
                            
                            all_images.append(img_info)
                            page_detail['image_bytes'] += img_size
                            result['total_image_bytes'] += img_size
                            
                            # Track colorspaces
                            result['image_colorspaces'][cs_name] = result['image_colorspaces'].get(cs_name, 0) + 1
                            
                            # Track max dimensions
                            result['max_image_width'] = max(result['max_image_width'], img_width)
                            result['max_image_height'] = max(result['max_image_height'], img_height)
                            pixels = img_width * img_height
                            result['max_image_pixels'] = max(result['max_image_pixels'], pixels)
                            
                            if pixels > 1000000:
                                result['images_over_1mp'] += 1
                            if pixels > 4000000:
                                result['images_over_4mp'] += 1
                            
                            # Store first few images per page
                            if len(page_detail['images']) < 3:
                                page_detail['images'].append(img_info)
                    except Exception as e:
                        pass
            except Exception as e:
                pass
            
            # Extract text
            try:
                text = page.get_text()
                page_detail['text_chars'] = len(text)
                result['total_text_chars'] += len(text)
            except:
                pass
            
            # Extract fonts
            try:
                fonts = page.get_fonts()
                page_detail['font_count'] = len(fonts)
                for font in fonts:
                    all_fonts.add(font[3] if len(font) > 3 else str(font))
            except:
                pass
            
            # Extract drawings (vector graphics)
            try:
                drawings = page.get_drawings()
                page_detail['drawing_count'] = len(drawings)
                result['total_drawings'] += len(drawings)
            except:
                pass
            
            # Extract links
            try:
                links = page.get_links()
                page_detail['link_count'] = len(links)
                result['total_links'] += len(links)
            except:
                pass
            
            # Extract annotations
            try:
                annots = list(page.annots()) if page.annots() else []
                page_detail['annot_count'] = len(annots)
                result['total_annotations'] += len(annots)
                if annots:
                    result['has_annotations'] = True
            except:
                pass
            
            # Store page details (limit to first 20 pages for Excel)
            if page_num < 20:
                result['page_details'].append(page_detail)
        
        doc.close()
        
        # Calculate aggregates
        result['total_fonts'] = len(all_fonts)
        result['unique_page_sizes'] = len(page_sizes_set)
        result['page_sizes'] = [list(ps) for ps in sorted(page_sizes_set)]
        
        if result['page_count'] > 0:
            result['mb_per_page'] = round(file_size_mb / result['page_count'], 3)
            result['images_per_page'] = round(result['total_images'] / result['page_count'], 2)
            result['chars_per_page'] = round(result['total_text_chars'] / result['page_count'], 0)
            result['drawings_per_page'] = round(result['total_drawings'] / result['page_count'], 1)
        
        if all_images:
            total_pixels = sum(img['pixels'] for img in all_images)
            result['avg_image_pixels'] = round(total_pixels / len(all_images), 0)
        
        # Risk assessment
        risk_factors = []
        complexity_score = 0
        
        # File size risks
        if file_size_mb > 100:
            risk_factors.append(f"File exceeds 100MB limit ({file_size_mb:.1f}MB)")
            complexity_score += 50
        elif file_size_mb > 50:
            risk_factors.append(f"Large file ({file_size_mb:.1f}MB)")
            complexity_score += 20
        
        # Page count risks
        if result['page_count'] > 200:
            risk_factors.append(f"Over 200 pages ({result['page_count']})")
            complexity_score += 30
        elif result['page_count'] > 100:
            risk_factors.append(f"Over 100 pages ({result['page_count']})")
            complexity_score += 15
        
        # MB per page (density)
        if result['mb_per_page'] > 5:
            risk_factors.append(f"Very high density ({result['mb_per_page']:.1f} MB/page)")
            complexity_score += 30
        elif result['mb_per_page'] > 2:
            risk_factors.append(f"High density ({result['mb_per_page']:.1f} MB/page)")
            complexity_score += 15
        
        # Large images
        if result['images_over_4mp'] > 0:
            risk_factors.append(f"{result['images_over_4mp']} images over 4MP")
            complexity_score += 20
        if result['images_over_1mp'] > 5:
            risk_factors.append(f"{result['images_over_1mp']} images over 1MP")
            complexity_score += 10
        
        # Page size issues
        if result['pages_over_tabloid'] > 0:
            risk_factors.append(f"{result['pages_over_tabloid']} pages larger than tabloid")
            complexity_score += 15
        
        if result['unique_page_sizes'] > 3:
            risk_factors.append(f"{result['unique_page_sizes']} different page sizes")
            complexity_score += 10
        
        # Image-heavy
        if result['images_per_page'] > 3:
            risk_factors.append(f"Image-heavy ({result['images_per_page']:.1f} images/page)")
            complexity_score += 15
        
        # CMYK images (can be problematic)
        cmyk_count = result['image_colorspaces'].get('CMYK', 0)
        if cmyk_count > 0:
            risk_factors.append(f"{cmyk_count} CMYK images")
            complexity_score += 5
        
        # Complex vector graphics
        if result['drawings_per_page'] > 100:
            risk_factors.append(f"Heavy vector graphics ({result['drawings_per_page']:.0f} drawings/page)")
            complexity_score += 20
        elif result['drawings_per_page'] > 50:
            risk_factors.append(f"Many vector graphics ({result['drawings_per_page']:.0f} drawings/page)")
            complexity_score += 10
        
        # JavaScript (can cause issues)
        if result['has_javascript']:
            risk_factors.append("Contains JavaScript")
            complexity_score += 10
        
        # Embedded files
        if result['has_embedded_files']:
            risk_factors.append("Contains embedded files")
            complexity_score += 5
        
        # Encrypted
        if result['is_encrypted']:
            risk_factors.append("Encrypted PDF")
            complexity_score += 25
        
        # Low text content (might be scanned)
        if result['page_count'] > 0 and result['chars_per_page'] < 100 and result['images_per_page'] > 0:
            risk_factors.append(f"Possibly scanned (only {result['chars_per_page']:.0f} chars/page)")
            complexity_score += 15
        
        result['risk_factors'] = risk_factors
        result['complexity_score'] = complexity_score
        
    except Exception as e:
        result['error'] = str(e)
        result['risk_factors'] = [f"Error analyzing PDF: {e}"]
        result['complexity_score'] = 100
    
    return result



def write_excel_report(results: list, output_path: str):
    """Write advanced analysis results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ===== Main Summary Sheet =====
    ws = wb.active
    ws.title = "Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    high_risk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    med_risk_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_risk_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        "Filename", "Pages", "Size (MB)", "MB/Page",
        "Images", "Img/Page", "Max Img MP", "Images >1MP", "Images >4MP",
        "Text Chars", "Chars/Page", "Fonts", "Drawings", "Draw/Page",
        "Page Sizes", "Over Tabloid", "Colorspaces",
        "Complexity Score", "Risk Factors"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    for row_num, r in enumerate(results, 2):
        colorspaces_str = ", ".join(f"{k}:{v}" for k, v in r['image_colorspaces'].items())
        risk_factors_str = "\n".join(r['risk_factors']) if r['risk_factors'] else ""
        max_img_mp = round(r['max_image_pixels'] / 1000000, 2) if r['max_image_pixels'] else 0
        
        row_data = [
            r['filename'],
            r['page_count'],
            r['file_size_mb'],
            r['mb_per_page'],
            r['total_images'],
            r['images_per_page'],
            max_img_mp,
            r['images_over_1mp'],
            r['images_over_4mp'],
            r['total_text_chars'],
            r['chars_per_page'],
            r['total_fonts'],
            r['total_drawings'],
            r['drawings_per_page'],
            r['unique_page_sizes'],
            r['pages_over_tabloid'],
            colorspaces_str,
            r['complexity_score'],
            risk_factors_str,
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Risk Factors column - wrap text
            if col == 19:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Complexity score coloring
            if col == 18:
                score = r['complexity_score']
                if score >= 50:
                    cell.fill = high_risk_fill
                elif score >= 25:
                    cell.fill = med_risk_fill
                else:
                    cell.fill = low_risk_fill
                cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    col_widths = [30, 8, 10, 10, 8, 10, 10, 10, 10, 12, 10, 8, 10, 10, 10, 12, 20, 12, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = 'A2'
    
    # ===== Image Details Sheet =====
    img_ws = wb.create_sheet("Image Details")
    
    img_headers = ["Filename", "Page", "Image #", "Width", "Height", "Megapixels", "Colorspace", "BPC", "Size (KB)"]
    for col, header in enumerate(img_headers, 1):
        cell = img_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    img_row = 2
    for r in results:
        for pd in r.get('page_details', []):
            for img_idx, img in enumerate(pd.get('images', []), 1):
                row_data = [
                    r['filename'],
                    pd['page'],
                    img_idx,
                    img['width'],
                    img['height'],
                    round(img['pixels'] / 1000000, 2),
                    img['colorspace'],
                    img['bpc'],
                    round(img['size_bytes'] / 1024, 1),
                ]
                for col, value in enumerate(row_data, 1):
                    cell = img_ws.cell(row=img_row, column=col, value=value)
                    cell.border = thin_border
                img_row += 1
    
    for col, width in enumerate([30, 8, 10, 10, 10, 12, 12, 8, 12], 1):
        img_ws.column_dimensions[get_column_letter(col)].width = width
    img_ws.freeze_panes = 'A2'
    
    # ===== Page Details Sheet =====
    page_ws = wb.create_sheet("Page Details")
    
    page_headers = ["Filename", "Page", "Width (pt)", "Height (pt)", "Rotation", 
                    "Images", "Image KB", "Text Chars", "Fonts", "Drawings", "Links", "Annotations"]
    for col, header in enumerate(page_headers, 1):
        cell = page_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    page_row = 2
    for r in results:
        for pd in r.get('page_details', []):
            row_data = [
                r['filename'],
                pd['page'],
                pd['width_pt'],
                pd['height_pt'],
                pd['rotation'],
                pd['image_count'],
                round(pd['image_bytes'] / 1024, 1),
                pd['text_chars'],
                pd['font_count'],
                pd['drawing_count'],
                pd['link_count'],
                pd['annot_count'],
            ]
            for col, value in enumerate(row_data, 1):
                cell = page_ws.cell(row=page_row, column=col, value=value)
                cell.border = thin_border
            page_row += 1
    
    for col, width in enumerate([30, 8, 12, 12, 10, 10, 12, 12, 8, 10, 8, 12], 1):
        page_ws.column_dimensions[get_column_letter(col)].width = width
    page_ws.freeze_panes = 'A2'
    
    # ===== Metadata Sheet =====
    meta_ws = wb.create_sheet("Metadata")
    
    meta_headers = ["Filename", "PDF Version", "Encrypted", "JavaScript", "Embedded Files",
                    "Title", "Author", "Creator", "Producer"]
    for col, header in enumerate(meta_headers, 1):
        cell = meta_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_num, r in enumerate(results, 2):
        meta = r.get('metadata', {})
        row_data = [
            r['filename'],
            r['pdf_version'],
            "Yes" if r['is_encrypted'] else "No",
            "Yes" if r['has_javascript'] else "No",
            "Yes" if r['has_embedded_files'] else "No",
            meta.get('title', ''),
            meta.get('author', ''),
            meta.get('creator', ''),
            meta.get('producer', ''),
        ]
        for col, value in enumerate(row_data, 1):
            cell = meta_ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    for col, width in enumerate([30, 15, 12, 12, 15, 30, 20, 25, 25], 1):
        meta_ws.column_dimensions[get_column_letter(col)].width = width
    meta_ws.freeze_panes = 'A2'
    
    # ===== Statistics Sheet =====
    stats_ws = wb.create_sheet("Statistics")
    
    total_files = len(results)
    high_risk = sum(1 for r in results if r['complexity_score'] >= 50)
    med_risk = sum(1 for r in results if 25 <= r['complexity_score'] < 50)
    low_risk = sum(1 for r in results if r['complexity_score'] < 25)
    total_pages = sum(r['page_count'] for r in results)
    total_size = sum(r['file_size_mb'] for r in results)
    total_images = sum(r['total_images'] for r in results)
    
    stats_data = [
        ["Advanced PDF Pre-Scan Statistics", ""],
        ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["Total PDFs", total_files],
        ["Total Pages", total_pages],
        ["Total Size (MB)", round(total_size, 2)],
        ["Total Images", total_images],
        ["", ""],
        ["Risk Distribution", ""],
        ["High Risk (score >= 50)", high_risk],
        ["Medium Risk (25-49)", med_risk],
        ["Low Risk (< 25)", low_risk],
        ["", ""],
        ["Averages", ""],
        ["Avg Pages/PDF", round(total_pages / total_files, 1) if total_files else 0],
        ["Avg Size/PDF (MB)", round(total_size / total_files, 2) if total_files else 0],
        ["Avg Images/PDF", round(total_images / total_files, 1) if total_files else 0],
    ]
    
    for row_num, row_data in enumerate(stats_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = stats_ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif str(value).endswith(('Distribution', 'Averages')):
                cell.font = Font(bold=True)
    
    stats_ws.column_dimensions['A'].width = 25
    stats_ws.column_dimensions['B'].width = 20
    
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")



def print_summary(result: dict):
    """Print a summary to console."""
    print(f"\n{'='*60}")
    print(f"PDF: {result['filename']}")
    print(f"{'='*60}")
    print(f"  Size: {result['file_size_mb']:.2f} MB | Pages: {result['page_count']} | {result['mb_per_page']:.2f} MB/page")
    print(f"  Images: {result['total_images']} ({result['images_per_page']:.1f}/page) | Max: {result['max_image_pixels']/1000000:.1f}MP")
    print(f"  Text: {result['total_text_chars']} chars | Fonts: {result['total_fonts']} | Drawings: {result['total_drawings']}")
    print(f"  Page sizes: {result['unique_page_sizes']} unique | Over tabloid: {result['pages_over_tabloid']}")
    print(f"  Colorspaces: {result['image_colorspaces']}")
    print()
    print(f"  Complexity Score: {result['complexity_score']}")
    if result['risk_factors']:
        print(f"  Risk Factors:")
        for rf in result['risk_factors']:
            print(f"    - {rf}")
    else:
        print(f"  Risk Factors: None")


def main():
    parser = argparse.ArgumentParser(
        description='Advanced PDF analysis using PyMuPDF with Excel export.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s document.pdf --xlsx report.xlsx
  %(prog)s /path/to/pdfs/ --xlsx analysis.xlsx
  %(prog)s /path/to/pdfs/ -r --xlsx recursive.xlsx
  %(prog)s document.pdf  # Console output only
        """
    )
    parser.add_argument('path', help='PDF file or directory containing PDFs')
    parser.add_argument('--xlsx', metavar='FILE', help='Export results to Excel spreadsheet')
    parser.add_argument('--recursive', '-r', action='store_true', help='Recursively scan directories')
    parser.add_argument('--quiet', '-q', action='store_true', help='Suppress console output')
    
    args = parser.parse_args()
    
    # Check for PyMuPDF
    try:
        import fitz
    except ImportError:
        print("Error: PyMuPDF is required. Install with: pip install pymupdf", file=sys.stderr)
        sys.exit(1)
    
    # Check for openpyxl if xlsx requested
    if args.xlsx:
        try:
            from openpyxl import Workbook
        except ImportError:
            print("Error: openpyxl is required for Excel export. Install with: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
    
    # Collect PDF files
    pdf_files = []
    path = Path(args.path)
    
    if path.is_file():
        if path.suffix.lower() == '.pdf':
            pdf_files.append(str(path))
        else:
            print(f"Error: {args.path} is not a PDF file", file=sys.stderr)
            sys.exit(1)
    elif path.is_dir():
        pattern = '**/*.pdf' if args.recursive else '*.pdf'
        pdf_files = [str(p) for p in path.glob(pattern)]
        if not pdf_files:
            print(f"No PDF files found in {args.path}", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"Error: {args.path} does not exist", file=sys.stderr)
        sys.exit(1)
    
    if not args.quiet:
        print(f"Analyzing {len(pdf_files)} PDF(s)...")
    
    # Analyze PDFs
    results = []
    for i, pdf_path in enumerate(sorted(pdf_files), 1):
        if not args.quiet:
            print(f"  [{i}/{len(pdf_files)}] {os.path.basename(pdf_path)}...", end=' ', flush=True)
        
        result = analyze_pdf_advanced(pdf_path)
        results.append(result)
        
        if not args.quiet:
            print(f"Score: {result['complexity_score']}")
    
    # Output
    if args.xlsx:
        write_excel_report(results, args.xlsx)
    
    if not args.quiet and not args.xlsx:
        for result in results:
            print_summary(result)
    
    # Print summary stats
    if not args.quiet:
        high_risk = sum(1 for r in results if r['complexity_score'] >= 50)
        med_risk = sum(1 for r in results if 25 <= r['complexity_score'] < 50)
        low_risk = sum(1 for r in results if r['complexity_score'] < 25)
        print(f"\nSummary: {len(results)} PDFs - {high_risk} high risk, {med_risk} medium, {low_risk} low")


if __name__ == '__main__':
    main()
