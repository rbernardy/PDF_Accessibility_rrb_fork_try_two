"""
PDF Processing Report Generator Lambda

This Lambda function generates an Excel report of all processed PDF files.
It retrieves file metadata and accessibility report data from S3.

Supports batch processing for large numbers of files:
- Processes files in batches to avoid Lambda timeout
- Saves intermediate results to S3
- Re-invokes itself for subsequent batches
- Merges all batch results into final Excel report

The report includes:
- File path, name, size, page count
- Before and after remediation accessibility metrics
"""

import os
import io
import json
import boto3
from datetime import datetime
from zoneinfo import ZoneInfo
from typing import Dict, List, Optional, Any
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Initialize clients
s3_client = boto3.client('s3')
lambda_client = boto3.client('lambda')

# Version marker to force Lambda updates
VERSION = "3.0.0-batch"

# Batch processing settings
BATCH_SIZE = 200  # Number of PDFs to process per invocation


def get_pdf_page_count(bucket: str, key: str) -> int:
    """
    Get the number of pages in a PDF file from S3.
    
    Args:
        bucket: S3 bucket name
        key: S3 object key
        
    Returns:
        Number of pages in the PDF, or 0 if unable to read
    """
    try:
        response = s3_client.get_object(Bucket=bucket, Key=key)
        pdf_bytes = response['Body'].read()
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return len(reader.pages)
    except Exception as e:
        print(f"Error getting page count for {key}: {e}")
        return 0


def get_file_size(bucket: str, key: str) -> int:
    """
    Get the size of a file in S3.
    
    Args:
        bucket: S3 bucket name
        key: S3 object key
        
    Returns:
        File size in bytes, or 0 if unable to get
    """
    try:
        response = s3_client.head_object(Bucket=bucket, Key=key)
        return response['ContentLength']
    except Exception as e:
        print(f"Error getting file size for {key}: {e}")
        return 0


def list_result_pdfs(bucket: str) -> List[Dict[str, Any]]:
    """
    List all PDF files in the result folder.
    
    Args:
        bucket: S3 bucket name
        
    Returns:
        List of dictionaries with file info (key, size, last_modified)
    """
    pdf_files = []
    paginator = s3_client.get_paginator('list_objects_v2')
    
    try:
        for page in paginator.paginate(Bucket=bucket, Prefix='result/'):
            for obj in page.get('Contents', []):
                key = obj['Key']
                if key.lower().endswith('.pdf'):
                    pdf_files.append({
                        'key': key,
                        'size': obj['Size'],
                        'last_modified': obj['LastModified'].isoformat()
                    })
    except Exception as e:
        print(f"Error listing PDFs in result folder: {e}")
    
    return pdf_files


def extract_folder_path_from_result_key(result_key: str) -> str:
    """
    Extract the folder path from a result key.
    
    Example: result/folder1/folder2/COMPLIANT_file.pdf -> folder1/folder2
    
    Args:
        result_key: The S3 key from the result folder
        
    Returns:
        The folder path without 'result/' prefix and filename
    """
    # Remove 'result/' prefix
    path_without_prefix = result_key.replace('result/', '', 1)
    # Remove the filename
    parts = path_without_prefix.rsplit('/', 1)
    if len(parts) > 1:
        return parts[0]
    return ''


def extract_original_filename(result_key: str) -> str:
    """
    Extract the original filename from a result key.
    
    Example: result/folder/COMPLIANT_file.pdf -> file.pdf
    
    Args:
        result_key: The S3 key from the result folder
        
    Returns:
        The original filename without COMPLIANT_ prefix
    """
    filename = os.path.basename(result_key)
    if filename.startswith('COMPLIANT_'):
        return filename.replace('COMPLIANT_', '', 1)
    return filename


def get_accessibility_report_path(folder_path: str, original_filename: str, report_type: str) -> str:
    """
    Construct the S3 path for an accessibility report.
    
    Args:
        folder_path: The folder path (e.g., 'folder1/folder2')
        original_filename: The original PDF filename without COMPLIANT_ prefix
        report_type: Either 'before' or 'after'
        
    Returns:
        The S3 key for the accessibility report
    """
    filename_without_ext = os.path.splitext(original_filename)[0]
    folder_prefix = f"{folder_path}/" if folder_path else ""
    
    if report_type == 'before':
        return f"temp/{folder_prefix}{filename_without_ext}/accessability-report/{filename_without_ext}_accessibility_report_before_remidiation.json"
    else:
        return f"temp/{folder_prefix}{filename_without_ext}/accessability-report/COMPLIANT_{filename_without_ext}_accessibility_report_after_remidiation.json"


def get_error_report_path(folder_path: str, original_filename: str) -> str:
    """
    Construct the S3 path for a pre-remediation error report.
    
    Args:
        folder_path: The folder path (e.g., 'folder1/folder2')
        original_filename: The original PDF filename without COMPLIANT_ prefix
        
    Returns:
        The S3 key for the error report
    """
    filename_without_ext = os.path.splitext(original_filename)[0]
    folder_prefix = f"{folder_path}/" if folder_path else ""
    return f"temp/{folder_prefix}{filename_without_ext}/accessability-report/{filename_without_ext}_pre_remediation_ERROR.json"


def load_json_from_s3(bucket: str, key: str) -> Optional[Dict]:
    """
    Load a JSON file from S3.
    
    Args:
        bucket: S3 bucket name
        key: S3 object key
        
    Returns:
        Parsed JSON as dictionary, or None if not found/error
    """
    try:
        response = s3_client.get_object(Bucket=bucket, Key=key)
        content = response['Body'].read().decode('utf-8')
        return json.loads(content)
    except s3_client.exceptions.NoSuchKey:
        print(f"JSON file not found: {key}")
        return None
    except Exception as e:
        print(f"Error loading JSON from {key}: {e}")
        return None


def flatten_json(data: Any, prefix: str = '') -> Dict[str, Any]:
    """
    Flatten a nested JSON structure into a flat dictionary.
    
    Handles the Adobe accessibility report structure:
    - summary array -> before-summary-description, before-summary-needs_manual_check, etc.
    - Detailed Report -> Document -> before-detailed_report-document-rule, etc.
    
    Args:
        data: The JSON data to flatten
        prefix: Prefix for keys (e.g., 'before' or 'after')
        
    Returns:
        Flattened dictionary with hierarchical column names
    """
    items = {}
    
    if data is None:
        return items
    
    if not isinstance(data, (dict, list)):
        # Simple value
        if prefix:
            items[prefix] = data
        return items
    
    if isinstance(data, list):
        # Handle arrays
        if not data:
            items[f"{prefix}-count"] = 0
            return items
        
        # Check if it's an array of objects (like summary array or Detailed Report)
        if isinstance(data[0], dict):
            # Flatten each object in the array
            # For arrays like summary, each item has description, status, etc.
            for i, item in enumerate(data):
                for key, value in item.items():
                    # Create column name: prefix-key (e.g., before-summary-description)
                    # For multiple items with same key, we'll use the first one or concatenate
                    col_name = f"{prefix}-{normalize_key(key)}"
                    
                    if isinstance(value, (dict, list)):
                        # Recursively flatten nested structures
                        nested = flatten_json(value, col_name)
                        items.update(nested)
                    else:
                        # For duplicate keys across array items, append index or concatenate
                        if col_name in items:
                            # Append with index for subsequent items
                            items[f"{col_name}-{i}"] = value
                        else:
                            items[col_name] = value
        else:
            # Array of simple values
            items[f"{prefix}-count"] = len(data)
            items[f"{prefix}-values"] = '; '.join(str(v) for v in data[:10])
        
        return items
    
    # Handle dictionaries
    for key, value in data.items():
        col_name = f"{prefix}-{normalize_key(key)}" if prefix else normalize_key(key)
        
        if isinstance(value, dict):
            # Recursively flatten nested dicts
            nested = flatten_json(value, col_name)
            items.update(nested)
        elif isinstance(value, list):
            # Handle arrays
            nested = flatten_json(value, col_name)
            items.update(nested)
        else:
            # Simple value
            items[col_name] = value
    
    return items


def normalize_key(key: str) -> str:
    """
    Normalize a JSON key to a column-friendly format.
    
    - Spaces become underscores
    - Convert to lowercase
    - Keep underscores as-is
    
    Args:
        key: The original key name
        
    Returns:
        Normalized key name
    """
    # Replace spaces with underscores, convert to lowercase
    normalized = str(key).replace(' ', '_').lower()
    return normalized


def build_report_row(bucket: str, pdf_info: Dict) -> Dict[str, Any]:
    """
    Build a single row of the report for a PDF file.
    
    Args:
        bucket: S3 bucket name
        pdf_info: Dictionary with PDF file info
        
    Returns:
        Dictionary representing a row in the CSV report
    """
    result_key = pdf_info['key']
    folder_path = extract_folder_path_from_result_key(result_key)
    original_filename = extract_original_filename(result_key)
    
    # Start with basic file info
    row = {
        'file-path': result_key,
        'file-name': os.path.basename(result_key),
        'original-filename': original_filename,
        'folder-path': folder_path,
        'file-size-bytes': pdf_info['size'],
        'last-modified': pdf_info['last_modified'],
        'page-count': get_pdf_page_count(bucket, result_key)
    }
    
    # Load before remediation report
    before_report_key = get_accessibility_report_path(folder_path, original_filename, 'before')
    before_data = load_json_from_s3(bucket, before_report_key)
    if before_data:
        row['before-report-found'] = True
        row['before-report-error'] = False
        flattened_before = flatten_json(before_data, 'before')
        row.update(flattened_before)
    else:
        row['before-report-found'] = False
        # Check if there's an error report
        error_report_key = get_error_report_path(folder_path, original_filename)
        error_data = load_json_from_s3(bucket, error_report_key)
        if error_data:
            row['before-report-error'] = True
            row['before-error-type'] = error_data.get('error_type', 'Unknown')
            row['before-error-message'] = error_data.get('error_message', 'Unknown error')
            row['before-error-timestamp'] = error_data.get('timestamp', '')
        else:
            row['before-report-error'] = False
            row['before-error-type'] = 'MissingReport'
            row['before-error-message'] = 'No before report or error log found'
    
    # Load after remediation report
    after_report_key = get_accessibility_report_path(folder_path, original_filename, 'after')
    after_data = load_json_from_s3(bucket, after_report_key)
    if after_data:
        row['after-report-found'] = True
        flattened_after = flatten_json(after_data, 'after')
        row.update(flattened_after)
    else:
        row['after-report-found'] = False
    
    return row


def collect_all_columns(rows: List[Dict]) -> List[str]:
    """
    Collect all unique column names from all rows.
    
    Orders columns as:
    1. Basic file info columns
    2. Status columns (report found, errors)
    3. All 'before' columns (in order they appear in JSON)
    4. All 'after' columns (in order they appear in JSON)
    
    Args:
        rows: List of row dictionaries
        
    Returns:
        Ordered list of all unique column names
    """
    # Use a list to preserve order of first appearance
    all_columns_ordered = []
    seen = set()
    
    for row in rows:
        for key in row.keys():
            if key not in seen:
                all_columns_ordered.append(key)
                seen.add(key)
    
    # Define column groups
    basic_cols = ['file-path', 'file-name', 'original-filename', 'folder-path', 
                  'file-size-bytes', 'last-modified', 'page-count']
    
    # Status columns - these come right after basic info
    status_cols = ['before-report-found', 'before-report-error', 'before-error-type', 
                   'before-error-message', 'before-error-timestamp', 'after-report-found']
    
    # Separate before and after columns (excluding status cols) - preserve order
    before_cols = [c for c in all_columns_ordered 
                   if c.startswith('before') and c not in status_cols]
    after_cols = [c for c in all_columns_ordered 
                  if c.startswith('after') and c not in status_cols]
    
    # Any other columns that don't fit the above categories - preserve order
    other_cols = [c for c in all_columns_ordered 
                  if c not in basic_cols 
                  and c not in status_cols
                  and not c.startswith('before') 
                  and not c.startswith('after')]
    
    # Filter status_cols to only include those that exist
    status_cols = [c for c in status_cols if c in seen]
    # Filter basic_cols to only include those that exist
    basic_cols = [c for c in basic_cols if c in seen]
    
    # Final order: basic -> status -> other -> before -> after
    return basic_cols + status_cols + other_cols + before_cols + after_cols


def generate_excel_content(rows: List[Dict], columns: List[str]) -> bytes:
    """
    Generate Excel content from rows and columns with formatting.
    
    Args:
        rows: List of row dictionaries
        columns: List of column names
        
    Returns:
        Excel file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Processing Report"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    # Write headers
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set header row height to accommodate wrapped text
    ws.row_dimensions[1].height = 60  # Taller height for wrapped headers
    
    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row.get(column, '')
            cell.value = value
            # Enable text wrapping for all cells
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths with reasonable limits
    for col_idx, column in enumerate(columns, start=1):
        # Set a fixed reasonable width that encourages wrapping
        # Basic file info columns can be wider, others narrower
        if column in ['file-path', 'file-name', 'original-filename']:
            ws.column_dimensions[get_column_letter(col_idx)].width = 40
        elif column in ['folder-path']:
            ws.column_dimensions[get_column_letter(col_idx)].width = 30
        elif column in ['file-size-bytes', 'page-count', 'last-modified']:
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        else:
            # All other columns (before/after data) - narrow to force wrapping
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def save_excel_to_s3(bucket: str, excel_content: bytes, suffix: str = '') -> str:
    """
    Save Excel content to S3.
    
    Args:
        bucket: S3 bucket name
        excel_content: Excel file content as bytes
        suffix: Optional suffix for the filename (e.g., '-batch-1')
        
    Returns:
        S3 key where the Excel file was saved
    """
    # Get timezone from environment variable, default to US/Eastern
    tz_name = os.environ.get('TZ', 'US/Eastern')
    try:
        tz = ZoneInfo(tz_name)
        timestamp = datetime.now(tz).strftime('%Y%m%d-%H%M%S')
    except Exception as e:
        print(f"Warning: Could not use timezone {tz_name}, falling back to UTC: {e}")
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    
    key = f"reports/pdf_processing_reports/pdf-processing-report-{timestamp}{suffix}.xlsx"
    
    s3_client.put_object(
        Bucket=bucket,
        Key=key,
        Body=excel_content,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    print(f"Excel report saved to s3://{bucket}/{key} (version: {VERSION})")
    return key


def save_batch_results_to_s3(bucket: str, rows: List[Dict], batch_id: str) -> str:
    """
    Save batch results as JSON to S3 for later merging.
    
    Args:
        bucket: S3 bucket name
        rows: List of row dictionaries
        batch_id: Unique identifier for this batch run
        
    Returns:
        S3 key where the batch results were saved
    """
    key = f"reports/pdf_processing_reports/temp/{batch_id}/batch-{len(rows)}.json"
    
    s3_client.put_object(
        Bucket=bucket,
        Key=key,
        Body=json.dumps(rows),
        ContentType='application/json'
    )
    
    print(f"Batch results saved to s3://{bucket}/{key}")
    return key


def load_batch_results_from_s3(bucket: str, batch_id: str) -> List[Dict]:
    """
    Load all batch results from S3.
    
    Args:
        bucket: S3 bucket name
        batch_id: Unique identifier for this batch run
        
    Returns:
        Combined list of all row dictionaries from all batches
    """
    all_rows = []
    prefix = f"reports/pdf_processing_reports/temp/{batch_id}/"
    
    paginator = s3_client.get_paginator('list_objects_v2')
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get('Contents', []):
            key = obj['Key']
            if key.endswith('.json'):
                try:
                    response = s3_client.get_object(Bucket=bucket, Key=key)
                    content = response['Body'].read().decode('utf-8')
                    rows = json.loads(content)
                    all_rows.extend(rows)
                    print(f"Loaded {len(rows)} rows from {key}")
                except Exception as e:
                    print(f"Error loading batch results from {key}: {e}")
    
    return all_rows


def cleanup_batch_results(bucket: str, batch_id: str):
    """
    Delete temporary batch result files from S3.
    
    Args:
        bucket: S3 bucket name
        batch_id: Unique identifier for this batch run
    """
    prefix = f"reports/pdf_processing_reports/temp/{batch_id}/"
    
    paginator = s3_client.get_paginator('list_objects_v2')
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get('Contents', []):
            try:
                s3_client.delete_object(Bucket=bucket, Key=obj['Key'])
                print(f"Deleted temp file: {obj['Key']}")
            except Exception as e:
                print(f"Error deleting {obj['Key']}: {e}")


def lambda_handler(event: Dict, context: Any) -> Dict:
    """
    Lambda handler for generating PDF processing reports.
    
    Supports batch processing:
    - First invocation: lists all PDFs, processes first batch, invokes next batch
    - Subsequent invocations: processes assigned batch, invokes next or finalizes
    
    Event parameters:
    - bucket: S3 bucket name (optional, defaults to BUCKET_NAME env var)
    - batch_id: Unique ID for this report generation run
    - batch_number: Current batch number (0-indexed)
    - total_files: Total number of PDF files to process
    - pdf_keys: List of PDF file keys to process in this batch (for continuation)
    
    Args:
        event: Lambda event
        context: Lambda context
        
    Returns:
        Response with status and report location
    """
    # Get bucket name from event or environment
    bucket = event.get('bucket') or os.environ.get('BUCKET_NAME')
    
    if not bucket:
        return {
            'statusCode': 400,
            'body': 'Missing bucket name. Provide in event or BUCKET_NAME env var.'
        }
    
    # Check if this is a continuation of batch processing
    batch_id = event.get('batch_id')
    batch_number = event.get('batch_number', 0)
    all_pdf_keys = event.get('all_pdf_keys')
    
    # First invocation - list all PDFs and start batch processing
    if not batch_id:
        print(f"Starting new report generation for bucket: {bucket} (version: {VERSION})")
        
        # Generate unique batch ID
        tz_name = os.environ.get('TZ', 'US/Eastern')
        try:
            tz = ZoneInfo(tz_name)
            batch_id = datetime.now(tz).strftime('%Y%m%d-%H%M%S')
        except:
            batch_id = datetime.now().strftime('%Y%m%d-%H%M%S')
        
        # List all PDFs in result folder
        pdf_files = list_result_pdfs(bucket)
        total_files = len(pdf_files)
        print(f"Found {total_files} PDF files in result folder")
        
        if not pdf_files:
            return {
                'statusCode': 200,
                'body': 'No PDF files found in result folder.'
            }
        
        # Store all PDF keys for batch processing
        all_pdf_keys = [f['key'] for f in pdf_files]
        
        # If small enough, process all at once (no batching needed)
        if total_files <= BATCH_SIZE:
            print(f"Processing all {total_files} files in single batch")
            rows = []
            for i, pdf_info in enumerate(pdf_files):
                print(f"Processing ({i+1}/{total_files}): {pdf_info['key']}")
                row = build_report_row(bucket, pdf_info)
                rows.append(row)
            
            columns = collect_all_columns(rows)
            excel_content = generate_excel_content(rows, columns)
            report_key = save_excel_to_s3(bucket, excel_content)
            
            return {
                'statusCode': 200,
                'body': {
                    'message': f'Report generated successfully with {len(rows)} files',
                    'report_location': f's3://{bucket}/{report_key}',
                    'files_processed': len(rows)
                }
            }
        
        # Large dataset - start batch processing
        print(f"Starting batch processing: {total_files} files in batches of {BATCH_SIZE}")
        batch_number = 0
    
    # Process current batch
    total_files = len(all_pdf_keys)
    start_idx = batch_number * BATCH_SIZE
    end_idx = min(start_idx + BATCH_SIZE, total_files)
    batch_keys = all_pdf_keys[start_idx:end_idx]
    
    print(f"Processing batch {batch_number + 1}: files {start_idx + 1} to {end_idx} of {total_files}")
    
    rows = []
    for i, key in enumerate(batch_keys):
        print(f"Processing ({start_idx + i + 1}/{total_files}): {key}")
        # Reconstruct pdf_info from key
        try:
            response = s3_client.head_object(Bucket=bucket, Key=key)
            pdf_info = {
                'key': key,
                'size': response['ContentLength'],
                'last_modified': response['LastModified'].isoformat()
            }
        except Exception as e:
            print(f"Error getting metadata for {key}: {e}")
            pdf_info = {'key': key, 'size': 0, 'last_modified': ''}
        
        row = build_report_row(bucket, pdf_info)
        rows.append(row)
    
    # Save batch results
    save_batch_results_to_s3(bucket, rows, batch_id)
    
    # Check if there are more batches
    if end_idx < total_files:
        # Invoke next batch
        next_batch = batch_number + 1
        print(f"Invoking next batch: {next_batch + 1}")
        
        function_name = context.function_name
        lambda_client.invoke(
            FunctionName=function_name,
            InvocationType='Event',  # Async invocation
            Payload=json.dumps({
                'bucket': bucket,
                'batch_id': batch_id,
                'batch_number': next_batch,
                'all_pdf_keys': all_pdf_keys
            })
        )
        
        return {
            'statusCode': 202,
            'body': {
                'message': f'Batch {batch_number + 1} complete, next batch triggered',
                'batch_id': batch_id,
                'files_processed_this_batch': len(rows),
                'total_files': total_files,
                'batches_remaining': (total_files - end_idx + BATCH_SIZE - 1) // BATCH_SIZE
            }
        }
    
    # Final batch - merge all results and generate Excel
    print("Final batch complete. Merging all results...")
    
    all_rows = load_batch_results_from_s3(bucket, batch_id)
    print(f"Loaded {len(all_rows)} total rows from all batches")
    
    columns = collect_all_columns(all_rows)
    excel_content = generate_excel_content(all_rows, columns)
    report_key = save_excel_to_s3(bucket, excel_content)
    
    # Cleanup temp files
    cleanup_batch_results(bucket, batch_id)
    
    return {
        'statusCode': 200,
        'body': {
            'message': f'Report generated successfully with {len(all_rows)} files',
            'report_location': f's3://{bucket}/{report_key}',
            'files_processed': len(all_rows),
            'batches_processed': batch_number + 1
        }
    }
