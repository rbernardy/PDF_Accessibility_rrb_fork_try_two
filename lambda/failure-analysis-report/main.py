"""
Failure Analysis Report Generator Lambda

Generates an Excel spreadsheet from PDF failure analysis data stored in DynamoDB.
Scheduled to run hourly via EventBridge.

Timing data (started_at, failed_at) is now stored permanently in the failure analysis
table at the moment of failure, so it doesn't depend on the in-flight tracker TTL.
Falls back to in-flight tracker for older entries that don't have timing data.
"""

import json
import os
import logging
import re
import boto3
from datetime import datetime, timezone, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger()
logger.setLevel(logging.INFO)

dynamodb = boto3.resource('dynamodb')
s3 = boto3.client('s3')

ANALYSIS_TABLE = os.environ.get('ANALYSIS_TABLE', '')
REPORT_BUCKET = os.environ.get('REPORT_BUCKET', '')
RATE_LIMIT_TABLE = os.environ.get('RATE_LIMIT_TABLE', 'adobe-api-in-flight-tracker')

# Prefix for individual file tracking entries (must match rate_limiter.py)
IN_FLIGHT_FILE_PREFIX = "file_"


def extract_collection_folder(s3_key: str) -> str:
    """
    Extract the collection folder from an S3 key.
    
    Example: pdf/my-collection-folder/document.pdf -> my-collection-folder
    Example: pdf/folder1/subfolder/document.pdf -> folder1/subfolder
    
    Args:
        s3_key: The S3 key of the PDF file
        
    Returns:
        The collection folder path, or empty string if not found
    """
    if not s3_key:
        return ''
    
    # Remove the 'pdf/' prefix if present
    if s3_key.startswith('pdf/'):
        path = s3_key[4:]
    else:
        path = s3_key
    
    # Split by '/' and remove the filename (last part)
    parts = path.rsplit('/', 1)
    if len(parts) > 1:
        return parts[0]
    return ''


def parse_adobe_error(original_error: str) -> dict:
    """
    Parse Adobe API error message to extract structured fields.
    
    Example error formats:
    - "ECS Task Failed (adobe-autotag): ... Adobe Autotag API failed: description =An Internal Server Error has occurred.;;"
    - "description =An Internal Server Error has occurred.;; requestTrackingId=...; statusCode=500; errorCode=INTERNAL_SERVER_ERROR"
    - "Status: 500"
    
    Returns:
        dict with 'api_name', 'description', 'status_code', 'error_code' keys
    """
    result = {
        'api_name': '',
        'description': '',
        'status_code': '',
        'error_code': ''
    }
    
    if not original_error:
        return result
    
    # Extract API name - try multiple patterns
    # Pattern 1: "Adobe Autotag API failed:" or "Adobe Extract API failed:"
    api_match = re.search(r'Adobe\s+(Autotag|Extract)\s+API\s+failed', original_error, re.IGNORECASE)
    if api_match:
        api_type = api_match.group(1).capitalize()
        result['api_name'] = f"Adobe {api_type} API"
    else:
        # Pattern 2: "ECS Task Failed (adobe-autotag)" or "ECS Task Failed (adobe-extract)"
        ecs_match = re.search(r'ECS\s+Task\s+Failed\s*\(\s*adobe[_-]?(autotag|extract)\s*\)', original_error, re.IGNORECASE)
        if ecs_match:
            api_type = ecs_match.group(1).capitalize()
            result['api_name'] = f"Adobe {api_type} API"
    
    # Try to extract description - multiple formats
    # Format 1: "description = ...;;" (with space around =, ends with ;;)
    desc_match = re.search(r"description\s*=\s*(.+?);;", original_error, re.IGNORECASE)
    if desc_match:
        result['description'] = desc_match.group(1).strip().strip("'\"")
    else:
        # Format 2: description='...' or description="..."
        desc_match = re.search(r"description\s*=\s*['\"]([^'\"]+)['\"]", original_error, re.IGNORECASE)
        if desc_match:
            result['description'] = desc_match.group(1)
        else:
            # Format 3: message=...;
            msg_match = re.search(r"message\s*=\s*([^;]+)", original_error, re.IGNORECASE)
            if msg_match:
                result['description'] = msg_match.group(1).strip()
    
    # Extract statusCode, httpStatusCode, or Status:
    status_match = re.search(r"(?:statusCode|httpStatusCode)\s*=\s*(\d+)", original_error, re.IGNORECASE)
    if status_match:
        result['status_code'] = status_match.group(1)
    else:
        # Try "Status: 500" format
        status_match = re.search(r"Status:\s*(\d+)", original_error, re.IGNORECASE)
        if status_match:
            result['status_code'] = status_match.group(1)
    
    # Extract errorCode
    error_code_match = re.search(r"errorCode\s*=\s*([A-Z_0-9]+)", original_error, re.IGNORECASE)
    if error_code_match:
        result['error_code'] = error_code_match.group(1)
    
    return result


def scan_all_items(table):
    """Scan all items from DynamoDB table with pagination."""
    items = []
    response = table.scan()
    items.extend(response.get('Items', []))
    
    while 'LastEvaluatedKey' in response:
        response = table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
        items.extend(response.get('Items', []))
    
    return items


def get_in_flight_timing_data(filename: str, api_type: str) -> dict:
    """
    Look up timing data from the adobe-api-in-flight-tracker table.
    
    Searches for file entries matching the filename and api_type.
    Returns the most recent entry's started_at and released_at timestamps.
    
    Note: released_at may be None if the container crashed before releasing the slot.
    In that case, the entry will have started_at but no released_at, indicating
    the process did not complete normally. However, if the ECS task failure tracker
    captured the crash, there will be a crashed_at timestamp instead.
    
    Args:
        filename: The PDF filename to look up
        api_type: The API type ('autotag' or 'extract')
        
    Returns:
        dict with 'started_at', 'released_at', 'crashed', 'crashed_at', and 'crash_details' keys
    """
    result = {
        'started_at': None,
        'released_at': None,
        'crashed': False,  # True if started_at exists but no released_at (container crash)
        'crashed_at': None,  # Timestamp from ECS task failure event
        'crash_details': None  # Details about the crash (exit code, reason, etc.)
    }
    
    try:
        table = dynamodb.Table(RATE_LIMIT_TABLE)
        
        # Scan for file entries matching this filename
        # Include both released and unreleased entries
        response = table.scan(
            FilterExpression='begins_with(counter_id, :prefix) AND filename = :filename',
            ExpressionAttributeValues={
                ':prefix': IN_FLIGHT_FILE_PREFIX,
                ':filename': filename
            }
        )
        
        items = response.get('Items', [])
        
        # Handle pagination
        while 'LastEvaluatedKey' in response:
            response = table.scan(
                FilterExpression='begins_with(counter_id, :prefix) AND filename = :filename',
                ExpressionAttributeValues={
                    ':prefix': IN_FLIGHT_FILE_PREFIX,
                    ':filename': filename
                },
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response.get('Items', []))
        
        if not items:
            logger.debug(f"No in-flight tracking data found for {filename}")
            return result
        
        # Filter by api_type if specified and available
        if api_type:
            matching_items = [item for item in items if item.get('api_type') == api_type]
            if matching_items:
                items = matching_items
        
        # Sort by started_at descending to get the most recent entry
        items.sort(key=lambda x: x.get('started_at', ''), reverse=True)
        
        # Get the most recent entry
        latest = items[0]
        result['started_at'] = latest.get('started_at')
        result['released_at'] = latest.get('released_at')
        result['crashed_at'] = latest.get('crashed_at')
        result['crash_details'] = latest.get('crash_details')
        
        # Determine if this was a crash
        # crashed_at is set by the ECS task failure tracker when a container crashes
        if result['crashed_at']:
            result['crashed'] = True
            # Use crashed_at as released_at if not already set
            if not result['released_at']:
                result['released_at'] = result['crashed_at']
        elif result['started_at'] and not result['released_at']:
            # No crashed_at but also no released_at - likely a crash that wasn't captured
            if not latest.get('released'):
                result['crashed'] = True
                logger.info(f"File {filename} appears to have crashed (started_at exists, no released_at)")
        
        return result
        
    except Exception as e:
        logger.warning(f"Error looking up in-flight timing data for {filename}: {e}")
        return result


def create_excel_report(items: list) -> bytes:
    """Create an Excel spreadsheet from the analysis data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Failure Analysis"
    
    # Define headers - includes parsed error fields and timing columns
    headers = [
        'Filename',
        'Collection Folder',
        'S3 Key',
        'Analysis Timestamp',
        'API Type',
        'Started At',
        'Released At',
        'Duration (s)',
        'Crashed',
        'Crash Reason',
        'Exit Code',
        'File Size (MB)',
        'Page Count',
        'Image Count',
        'Font Count',
        'Encrypted',
        'PDF Version',
        'Likely Cause',
        'Issue Count',
        'Error API',
        'Error Description',
        'Error Status Code',
        'Error Code',
        'Original Error'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    crash_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_num, item in enumerate(items, 2):
        # Calculate duration if both timestamps exist
        duration = ''
        started_at = item.get('started_at', '')
        # Use failed_at if available (new approach), otherwise use released_at
        released_at = item.get('failed_at', '') or item.get('released_at', '')
        crashed = item.get('crashed', False)
        
        if started_at and released_at:
            try:
                start_dt = datetime.fromisoformat(started_at.replace('Z', '+00:00'))
                end_dt = datetime.fromisoformat(released_at.replace('Z', '+00:00'))
                duration = round((end_dt - start_dt).total_seconds(), 1)
            except Exception:
                duration = ''
        
        # Extract crash details if available
        crash_reason = ''
        exit_code = ''
        crash_details = item.get('crash_details')
        if crash_details:
            try:
                if isinstance(crash_details, str):
                    details = json.loads(crash_details)
                else:
                    details = crash_details
                crash_reason = details.get('stopped_reason', '') or details.get('container_reason', '')
                exit_code = details.get('exit_code', '')
            except Exception:
                pass
        
        # Parse Adobe error message for structured fields
        original_error = item.get('original_error', '')
        parsed_error = parse_adobe_error(original_error)
        
        # Extract collection folder from S3 key
        s3_key = item.get('s3_key', '')
        collection_folder = extract_collection_folder(s3_key)
        
        row_data = [
            item.get('filename', ''),
            collection_folder,
            s3_key,
            item.get('analysis_timestamp', ''),
            item.get('api_type', ''),
            started_at,
            released_at,
            duration,
            'Yes' if crashed else 'No',
            crash_reason[:100] if crash_reason else '',  # Truncate long reasons
            exit_code if exit_code is not None else '',
            item.get('file_size_mb', ''),
            item.get('page_count', 0),
            item.get('image_count', 0),
            item.get('font_count', 0),
            'Yes' if item.get('has_encryption') else 'No',
            item.get('pdf_version', ''),
            item.get('likely_cause', ''),
            item.get('issue_count', 0),
            parsed_error['api_name'],
            parsed_error['description'],
            parsed_error['status_code'],
            parsed_error['error_code'],
            original_error[:500]  # Truncate long errors
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            # Highlight crashed rows
            if crashed:
                cell.fill = crash_fill
    
    # Auto-adjust column widths - updated for new columns
    column_widths = [30, 30, 50, 25, 12, 25, 25, 12, 10, 40, 10, 12, 12, 12, 12, 10, 12, 50, 12, 20, 40, 12, 25, 60]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add Excel table with auto-filter enabled
    if len(items) > 0:
        # Define table range (A1 to last column/row)
        last_col_letter = get_column_letter(len(headers))
        last_row = len(items) + 1  # +1 for header row
        table_ref = f"A1:{last_col_letter}{last_row}"
        
        # Create table with filters
        table = Table(displayName="FailureAnalysis", ref=table_ref)
        
        # Add a default table style with filters enabled
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Unprotect the sheet to allow editing (sheets are editable by default, but explicitly ensure it)
    ws.protection.sheet = False
    ws.protection.password = None
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def handler(event, context):
    """Generate Excel report from failure analysis data."""
    logger.info(f"Received event: {json.dumps(event)}")
    
    if not ANALYSIS_TABLE:
        logger.error("ANALYSIS_TABLE not configured")
        return {'statusCode': 500, 'body': 'ANALYSIS_TABLE not configured'}
    
    if not REPORT_BUCKET:
        logger.error("REPORT_BUCKET not configured")
        return {'statusCode': 500, 'body': 'REPORT_BUCKET not configured'}
    
    # Scan all items from DynamoDB
    table = dynamodb.Table(ANALYSIS_TABLE)
    logger.info(f"Scanning table: {ANALYSIS_TABLE}")
    items = scan_all_items(table)
    logger.info(f"Found {len(items)} analysis records")
    
    if not items:
        logger.info("No analysis data found, skipping report generation")
        return {
            'statusCode': 200,
            'body': json.dumps({'message': 'No analysis data found'})
        }
    
    # Enrich items with timing data
    # Priority: 1) Use timing data stored in failure analysis table (permanent)
    #           2) Fall back to in-flight tracker (may have expired)
    logger.info(f"Enriching items with timing data")
    for item in items:
        filename = item.get('filename', '')
        api_type = item.get('api_type', '')
        
        # Check if timing data is already in the failure analysis record (new approach)
        started_at = item.get('started_at', '')
        failed_at = item.get('failed_at', '')
        
        if started_at and failed_at:
            # Use timing data from failure analysis table (permanent storage)
            item['released_at'] = failed_at  # Use failed_at as released_at for consistency
            item['crashed'] = True  # If we have failure analysis, it was a failure
            logger.debug(f"Using timing data from failure analysis table for {filename}")
        elif filename:
            # Fall back to in-flight tracker for older entries
            timing_data = get_in_flight_timing_data(filename, api_type)
            if not started_at:
                item['started_at'] = timing_data.get('started_at', '')
            if not item.get('released_at'):
                item['released_at'] = timing_data.get('released_at', '')
            item['crashed'] = timing_data.get('crashed', False)
            item['crashed_at'] = timing_data.get('crashed_at', '')
            item['crash_details'] = timing_data.get('crash_details')
    
    # Sort by timestamp descending
    items.sort(key=lambda x: x.get('analysis_timestamp', ''), reverse=True)
    
    # Generate Excel report
    excel_bytes = create_excel_report(items)
    
    # Generate filename with timestamp (US Eastern time)
    eastern_offset = timedelta(hours=-5)
    now_eastern = datetime.now(timezone.utc) + eastern_offset
    timestamp_str = now_eastern.strftime('%Y%m%d_%H%M%S')
    report_key = f"reports/failure_analysis_summary/failure_analysis_report_{timestamp_str}.xlsx"
    
    # Upload to S3
    s3.put_object(
        Bucket=REPORT_BUCKET,
        Key=report_key,
        Body=excel_bytes,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Count crashes for summary
    crash_count = sum(1 for item in items if item.get('crashed', False))
    
    logger.info(f"Saved report to s3://{REPORT_BUCKET}/{report_key}")
    logger.info(f"Report contains {len(items)} records, {crash_count} with container crashes")
    
    return {
        'statusCode': 200,
        'body': json.dumps({
            'message': 'Report generated successfully',
            'record_count': len(items),
            'crash_count': crash_count,
            'report_location': f"s3://{REPORT_BUCKET}/{report_key}"
        })
    }
